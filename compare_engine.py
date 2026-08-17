"""
Core comparison engine for the Seller-Input vs AI-Enrichment QC tool.

Given two dataframes (input / output), a shared key column, and a list of
common attribute columns to compare, this module:
  - normalizes values for a fair comparison (trims whitespace, treats
    blank/NaN as equivalent, optional case-insensitive compare)
  - builds a merged row-per-SKU comparison table
  - backfills AI Output values from Input when AI Output is blank
  - builds a summary of how many SKUs deviate per attribute, split by
    deviation type (value mismatch vs. missing-in-output/backfilled)
  - writes a formatted, highlighted .xlsx report with a color legend
"""

import re
import io
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IMAGE_URL_PATTERN = re.compile(r"(image|photo|img|url|link)", re.IGNORECASE)

NAVY = "1F2A44"
ORANGE = "E8720C"
YELLOW = "FFF2AC"        # value mismatch (both present, differ)
AMBER = "FFC97A"         # missing in AI output -> backfilled from Input
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"

# deviation categories
MATCH = ""
MISMATCH = "mismatch"
MISSING_OUTPUT = "missing_output"


def normalize_key(value):
    """Normalize a key value (styleId etc.) so '39700559' == 39700559 == 39700559.0"""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except ValueError:
        return s


def is_blank(value):
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    s = str(value).strip()
    return s == "" or s.lower() == "nan"


def normalize_value(value, ignore_case_whitespace=True):
    """Normalize a cell value for comparison purposes."""
    if is_blank(value):
        return ""
    s = str(value).strip()
    if ignore_case_whitespace:
        s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def detect_image_url_columns(columns):
    return [c for c in columns if IMAGE_URL_PATTERN.search(str(c))]


def detect_key_column(input_cols, output_cols):
    common = [c for c in input_cols if c in output_cols]
    priority = ["styleId", "StyleId", "SKU Id", "SKU", "sku_id", "Style Id", "styleid"]
    for p in priority:
        if p in common:
            return p
    for c in common:
        if re.search(r"(style\s*id|sku)", str(c), re.IGNORECASE):
            return c
    return common[0] if common else None


def common_attribute_columns(input_cols, output_cols, key_col, excluded_cols):
    common = [c for c in input_cols if c in output_cols]
    return [c for c in common if c != key_col and c not in excluded_cols]


@dataclass
class ComparisonResult:
    merged_df: pd.DataFrame
    attribute_cols: list
    key_col: str
    excluded_cols: list
    input_only_keys: list
    output_only_keys: list
    total_skus: int
    total_attributes: int
    skus_with_deviation: int
    skus_with_backfill: int
    attribute_summary: pd.DataFrame  # Attribute | Value Mismatch | Missing in Output | Total Deviating | % Deviating
    deviation_matrix: pd.DataFrame   # category strings, index=key, columns=attribute_cols
    input_columns: list              # full original Input file column order
    output_columns: list             # full original Output file column order
    input_matched_df: pd.DataFrame   # full original Input columns, matched rows only, indexed by _key
    output_matched_df: pd.DataFrame  # full original Output columns, matched rows only, backfilled, indexed by _key


def run_comparison(
    input_df: pd.DataFrame,
    output_df: pd.DataFrame,
    key_col: str,
    attribute_cols: list,
    excluded_cols: list,
    ignore_case_whitespace: bool = True,
) -> ComparisonResult:
    in_df = input_df.copy()
    out_df = output_df.copy()

    in_df["_key"] = in_df[key_col].apply(normalize_key)
    out_df["_key"] = out_df[key_col].apply(normalize_key)

    in_df = in_df[in_df["_key"].notna()].drop_duplicates(subset="_key", keep="first")
    out_df = out_df[out_df["_key"].notna()].drop_duplicates(subset="_key", keep="first")

    input_keys = set(in_df["_key"])
    output_keys = set(out_df["_key"])
    matched_keys = sorted(input_keys & output_keys, key=lambda x: (len(x), x))
    input_only = sorted(input_keys - output_keys, key=lambda x: (len(x), x))
    output_only = sorted(output_keys - input_keys, key=lambda x: (len(x), x))

    in_idx = in_df.set_index("_key")
    out_idx = out_df.set_index("_key")

    rows = []
    deviation_rows = []
    for k in matched_keys:
        row = {"_key": k, key_col: in_idx.loc[k, key_col]}
        dev_row = {"_key": k}
        for attr in attribute_cols:
            in_val = in_idx.loc[k, attr] if attr in in_idx.columns else None
            out_val = out_idx.loc[k, attr] if attr in out_idx.columns else None

            norm_in = normalize_value(in_val, ignore_case_whitespace)
            norm_out = normalize_value(out_val, ignore_case_whitespace)

            if norm_in == norm_out:
                category = MATCH
                display_out = out_val
            elif is_blank(out_val) and not is_blank(in_val):
                # AI Output missing this field entirely -> backfill from Input
                category = MISSING_OUTPUT
                display_out = in_val
            else:
                category = MISMATCH
                display_out = out_val

            row[f"{attr} :: Input"] = in_val
            row[f"{attr} :: AI Output"] = display_out
            dev_row[attr] = category
        rows.append(row)
        deviation_rows.append(dev_row)

    merged_df = pd.DataFrame(rows)
    deviation_matrix = pd.DataFrame(deviation_rows).set_index("_key") if deviation_rows else pd.DataFrame()

    # ---- Full original-structure dataframes (all columns, matched rows only) ----
    input_columns = list(input_df.columns)
    output_columns = list(output_df.columns)

    input_matched_df = in_idx.loc[matched_keys, input_columns] if matched_keys else in_idx.loc[[], input_columns]
    output_matched_df = out_idx.loc[matched_keys, output_columns].astype(object).copy() if matched_keys else out_idx.loc[[], output_columns].astype(object)

    # Backfill only the comparable attribute columns in the Output copy; every
    # other original Output column (workflow fields, images, etc.) is left untouched.
    for attr in attribute_cols:
        if attr not in output_matched_df.columns:
            continue
        for k in matched_keys:
            category = deviation_matrix.loc[k, attr] if (not deviation_matrix.empty and attr in deviation_matrix.columns) else MATCH
            if category == MISSING_OUTPUT:
                output_matched_df.loc[k, attr] = in_idx.loc[k, attr]

    total_skus = len(matched_keys)
    total_attributes = len(attribute_cols)

    if not deviation_matrix.empty:
        is_dev = deviation_matrix[attribute_cols].isin([MISMATCH, MISSING_OUTPUT])
        is_backfill = deviation_matrix[attribute_cols] == MISSING_OUTPUT
        is_mismatch = deviation_matrix[attribute_cols] == MISMATCH

        skus_with_deviation = int(is_dev.any(axis=1).sum())
        skus_with_backfill = int(is_backfill.any(axis=1).sum())

        mismatch_counts = is_mismatch.sum(axis=0)
        backfill_counts = is_backfill.sum(axis=0)
        total_counts = mismatch_counts + backfill_counts
    else:
        skus_with_deviation = 0
        skus_with_backfill = 0
        mismatch_counts = pd.Series({a: 0 for a in attribute_cols})
        backfill_counts = pd.Series({a: 0 for a in attribute_cols})
        total_counts = pd.Series({a: 0 for a in attribute_cols})

    attribute_summary = pd.DataFrame({
        "Attribute": total_counts.index,
        "Value Mismatch": mismatch_counts.values,
        "Missing in AI Output": backfill_counts.values,
        "Total Deviating": total_counts.values,
    })
    attribute_summary["% Deviating"] = (
        attribute_summary["Total Deviating"] / total_skus * 100 if total_skus else 0
    )
    attribute_summary = attribute_summary.sort_values(
        "Total Deviating", ascending=False
    ).reset_index(drop=True)

    return ComparisonResult(
        merged_df=merged_df,
        attribute_cols=attribute_cols,
        key_col=key_col,
        excluded_cols=excluded_cols,
        input_only_keys=input_only,
        output_only_keys=output_only,
        total_skus=total_skus,
        total_attributes=total_attributes,
        skus_with_deviation=skus_with_deviation,
        skus_with_backfill=skus_with_backfill,
        attribute_summary=attribute_summary,
        deviation_matrix=deviation_matrix,
        input_columns=input_columns,
        output_columns=output_columns,
        input_matched_df=input_matched_df,
        output_matched_df=output_matched_df,
    )


def build_report_workbook(result: ComparisonResult) -> bytes:
    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws_sum = wb.active
    ws_sum.title = "Summary"

    title_font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    label_font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    normal_font = Font(name="Calibri", size=11, color="000000")
    metric_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    amber_fill = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")

    ws_sum["B2"] = "Seller Input vs AI Enrichment — QC Comparison Summary"
    ws_sum["B2"].font = title_font
    ws_sum.merge_cells("B2:E2")

    # ---- Legend ----
    r = 3
    legend_start = r + 1
    ws_sum.cell(row=legend_start, column=2, value="Legend:").font = label_font
    ws_sum.cell(row=legend_start + 1, column=2).fill = yellow_fill
    ws_sum.cell(row=legend_start + 1, column=2).border = border
    ws_sum.cell(row=legend_start + 1, column=3, value="Value Mismatch — both Input and AI Output have data, but the values differ. Review and correct.").font = normal_font
    ws_sum.cell(row=legend_start + 2, column=2).fill = amber_fill
    ws_sum.cell(row=legend_start + 2, column=2).border = border
    ws_sum.cell(row=legend_start + 2, column=3, value="Missing in AI Output — AI did not generate this field. Input value has been auto-filled here for reference; confirm or replace.").font = normal_font

    metrics_start = legend_start + 4
    metrics = [
        ("SKUs Compared", result.total_skus),
        ("Attributes Compared", result.total_attributes),
        ("SKUs with at Least 1 Deviation", result.skus_with_deviation),
        ("SKUs with at Least 1 Backfilled Field", result.skus_with_backfill),
        ("Attributes with at Least 1 Deviation",
         int((result.attribute_summary["Total Deviating"] > 0).sum())),
        ("SKUs Found Only in Input File", len(result.input_only_keys)),
        ("SKUs Found Only in AI Output File", len(result.output_only_keys)),
    ]

    r = metrics_start
    for label, value in metrics:
        c_label = ws_sum.cell(row=r, column=2, value=label)
        c_label.font = label_font
        c_label.fill = metric_fill
        c_label.border = border
        c_val = ws_sum.cell(row=r, column=4, value=value)
        c_val.font = Font(name="Calibri", size=12, bold=True, color=ORANGE)
        c_val.border = border
        c_val.alignment = Alignment(horizontal="center")
        r += 1

    r += 2
    ws_sum.cell(row=r, column=2, value="Attribute-Level Deviation Breakdown").font = Font(
        name="Calibri", size=13, bold=True, color=NAVY
    )
    r += 1
    header_row = r
    headers = ["Attribute", "Value Mismatch", "Missing in AI Output", "Total Deviating", "% Deviating"]
    for ci, h in enumerate(headers, start=2):
        cell = ws_sum.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1
    for _, row_data in result.attribute_summary.iterrows():
        ws_sum.cell(row=r, column=2, value=row_data["Attribute"]).font = normal_font
        for ci, key in zip((3, 4, 5), ("Value Mismatch", "Missing in AI Output", "Total Deviating")):
            c = ws_sum.cell(row=r, column=ci, value=int(row_data[key]))
            c.alignment = Alignment(horizontal="center")
            c.font = normal_font
        pct_cell = ws_sum.cell(row=r, column=6, value=round(float(row_data["% Deviating"]), 1))
        pct_cell.alignment = Alignment(horizontal="center")
        pct_cell.font = normal_font
        pct_cell.number_format = '0.0"%"'

        if row_data["Missing in AI Output"] > 0:
            fill = amber_fill
        elif row_data["Value Mismatch"] > 0:
            fill = yellow_fill
        else:
            fill = None
        if fill:
            for ci in (2, 3, 4, 5, 6):
                ws_sum.cell(row=r, column=ci).fill = fill
        for ci in (2, 3, 4, 5, 6):
            ws_sum.cell(row=r, column=ci).border = border
        r += 1

    ws_sum.column_dimensions["A"].width = 2
    ws_sum.column_dimensions["B"].width = 42
    ws_sum.column_dimensions["C"].width = 15
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 15
    ws_sum.column_dimensions["F"].width = 13

    if result.input_only_keys or result.output_only_keys:
        r += 2
        ws_sum.cell(row=r, column=2, value="Note: Unmatched SKUs (excluded from comparison)").font = Font(
            name="Calibri", size=11, bold=True, color=NAVY
        )
        r += 1
        if result.input_only_keys:
            ws_sum.cell(row=r, column=2, value=f"Only in Input file ({len(result.input_only_keys)}):").font = label_font
            ws_sum.cell(row=r, column=3, value=", ".join(result.input_only_keys[:50])).font = normal_font
            r += 1
        if result.output_only_keys:
            ws_sum.cell(row=r, column=2, value=f"Only in AI Output file ({len(result.output_only_keys)}):").font = label_font
            ws_sum.cell(row=r, column=3, value=", ".join(result.output_only_keys[:50])).font = normal_font
            r += 1

    # ---------- Sheet 2 & 3: Input Data / AI Output Data (full original structure) ----------
    key_col = result.key_col
    attrs = set(result.attribute_cols)
    dev_matrix = result.deviation_matrix

    def write_full_sheet(sheet_name, columns, matched_df, highlight_fn):
        ws = wb.create_sheet(sheet_name)

        for ci, col_name in enumerate(columns, start=1):
            c = ws.cell(row=1, column=ci, value=col_name)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_i = 2
        for key in matched_df.index:
            for ci, col_name in enumerate(columns, start=1):
                val = matched_df.loc[key, col_name] if col_name in matched_df.columns else None
                cell = ws.cell(row=row_i, column=ci, value=val if pd.notna(val) else "")
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if col_name in attrs:
                    category = dev_matrix.loc[key, col_name] if (not dev_matrix.empty and key in dev_matrix.index and col_name in dev_matrix.columns) else MATCH
                    fill = highlight_fn(category)
                    if fill:
                        cell.fill = fill
            row_i += 1

        for ci in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 22
        ws.freeze_panes = "B2"
        return ws

    # Input Data sheet: exact original Input file columns/values for matched SKUs.
    # Highlight cells that mismatch the AI Output (a field that's simply missing in
    # AI Output isn't the Input's problem, so no highlight there).
    write_full_sheet(
        "Input Data",
        result.input_columns,
        result.input_matched_df,
        lambda category: yellow_fill if category == MISMATCH else None,
    )

    # AI Output Data sheet: exact original Output file columns/values for matched
    # SKUs (workflow columns like Actions/AM/Detected Taxonomy pass through
    # untouched). Comparable attributes are highlighted: yellow for mismatches,
    # amber for cells that were blank and got backfilled from Input.
    write_full_sheet(
        "AI Output Data",
        result.output_columns,
        result.output_matched_df,
        lambda category: yellow_fill if category == MISMATCH else (amber_fill if category == MISSING_OUTPUT else None),
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
