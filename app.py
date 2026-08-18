import io
import pandas as pd
import streamlit as st

from compare_engine import (
    detect_key_column,
    detect_image_url_columns,
    common_attribute_columns,
    run_comparison,
    build_report_workbook,
)
from content_generator import (
    CATEGORIES,
    get_required_fields,
    guess_column,
    generate_all,
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

st.set_page_config(page_title="Rubick QC & Content Tools", layout="wide")

NAVY = "#1F2A44"
ORANGE = "#E8720C"
GREEN_FILL = "C6EFCE"

mode = st.sidebar.radio("Choose a tool", ["QC Comparison", "Generate Content"])

TARGET_COLUMNS = {
    "DN": "productDisplayName",
    "LVN": "listViewName",
    "PD": "Product Details",
    "Size & Fit": "sizeAndFitDescription",
}


def load_sheets(uploaded_file):
    """Return dict of {sheet_name: DataFrame} for an uploaded xlsx/csv."""
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
        return {"Sheet1": df}
    xls = pd.ExcelFile(uploaded_file)
    return {sn: xls.parse(sn) for sn in xls.sheet_names}


def pick_sheet(label, sheets_dict, key_prefix):
    """UI for picking which sheet holds the actual SKU data."""
    sheet_names = list(sheets_dict.keys())
    if len(sheet_names) == 1:
        chosen = sheet_names[0]
    else:
        info = [f"{sn} ({sheets_dict[sn].shape[0]} rows x {sheets_dict[sn].shape[1]} cols)" for sn in sheet_names]
        default_idx = max(range(len(sheet_names)), key=lambda i: sheets_dict[sheet_names[i]].shape[1])
        display_choice = st.selectbox(
            f"{label} — multiple sheets found, choose the one with SKU data:",
            options=info,
            index=default_idx,
            key=f"{key_prefix}_sheet",
        )
        chosen = sheet_names[info.index(display_choice)]
    return chosen, sheets_dict[chosen]


if mode == "QC Comparison":
    st.markdown(
        f"""
        <div style="background-color:{NAVY};padding:18px 24px;border-radius:6px;margin-bottom:14px;">
            <h2 style="color:white;margin:0;">Seller Input vs AI Enrichment — QC Comparison Tool</h2>
            <p style="color:#D6DCE5;margin:4px 0 0 0;">
                Upload the Seller/Vendor input file and the AI-enriched output file.
                The tool matches SKUs, compares every shared attribute, and highlights
                deviations for manual review. If a field is missing in the AI Output,
                the Input value is copied in automatically so no data is lost.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div style="display:flex;gap:24px;margin-bottom:20px;font-size:14px;">
            <div style="display:flex;align-items:center;gap:8px;">
                <span style="width:16px;height:16px;background-color:#FFF2AC;border:1px solid #ccc;display:inline-block;"></span>
                <span><b>Value Mismatch</b> — both files have data, but values differ. Review and correct.</span>
            </div>
            <div style="display:flex;align-items:center;gap:8px;">
                <span style="width:16px;height:16px;background-color:#FFC97A;border:1px solid #ccc;display:inline-block;"></span>
                <span><b>Missing in AI Output</b> — AI didn't generate this field; Input value auto-filled here. Confirm or replace.</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        input_file = st.file_uploader("1. Seller / Vendor Input file", type=["xlsx", "xls", "csv"], key="input_upl")
    with col2:
        output_file = st.file_uploader("2. AI Enrichment Output file", type=["xlsx", "xls", "csv"], key="output_upl")

    if input_file and output_file:
        input_sheets = load_sheets(input_file)
        output_sheets = load_sheets(output_file)

        st.markdown("### Select data sheets")
        scol1, scol2 = st.columns(2)
        with scol1:
            in_sheet_name, input_df = pick_sheet("Input file", input_sheets, "in")
        with scol2:
            out_sheet_name, output_df = pick_sheet("Output file", output_sheets, "out")

        input_df.columns = [str(c).strip() for c in input_df.columns]
        output_df.columns = [str(c).strip() for c in output_df.columns]

        default_key = detect_key_column(list(input_df.columns), list(output_df.columns))
        common_cols = [c for c in input_df.columns if c in list(output_df.columns)]

        if not common_cols:
            st.error("No matching column headers found between the two files. Please check the sheets selected above.")
            st.stop()

        st.markdown("### Matching key")
        key_col = st.selectbox(
            "Column used to match SKUs between the two files:",
            options=common_cols,
            index=common_cols.index(default_key) if default_key in common_cols else 0,
        )

        detected_image_url = detect_image_url_columns([c for c in common_cols if c != key_col])
        excluded_cols = st.multiselect(
            f"Exclude specific columns from comparison (optional — {len(detected_image_url)} image/URL "
            "columns detected, none excluded by default; add any here you don't want compared/highlighted):",
            options=[c for c in common_cols if c != key_col],
            default=[],
        )

        attribute_cols = common_attribute_columns(
            list(input_df.columns), list(output_df.columns), key_col, excluded_cols
        )

        ignore_case_whitespace = st.checkbox(
            "Ignore case & extra whitespace differences (recommended)", value=True
        )

        st.caption(
            f"Comparing **{len(attribute_cols)}** shared attributes across matched SKUs. "
            f"Key: **{key_col}**. Input sheet: *{in_sheet_name}* ({input_df.shape[0]} rows). "
            f"Output sheet: *{out_sheet_name}* ({output_df.shape[0]} rows)."
        )

        if st.button("Run Comparison", type="primary"):
            with st.spinner("Comparing SKUs and building report..."):
                result = run_comparison(
                    input_df, output_df, key_col, attribute_cols, excluded_cols, ignore_case_whitespace
                )
                report_bytes = build_report_workbook(result)

            st.success("Comparison complete.")

            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("SKUs Compared", result.total_skus)
            m2.metric("Attributes Compared", result.total_attributes)
            m3.metric("SKUs with ≥1 Deviation", result.skus_with_deviation)
            m4.metric("SKUs with ≥1 Backfilled Field", result.skus_with_backfill)
            attrs_with_dev = int((result.attribute_summary["Total Deviating"] > 0).sum())
            m5.metric("Attributes with ≥1 Deviation", attrs_with_dev)

            if result.input_only_keys or result.output_only_keys:
                st.warning(
                    f"{len(result.input_only_keys)} SKU(s) found only in the Input file, "
                    f"{len(result.output_only_keys)} SKU(s) found only in the Output file. "
                    "These were excluded from the comparison."
                )

            st.markdown("#### Attribute-level deviation breakdown")
            display_df = result.attribute_summary.copy()
            display_df["% Deviating"] = display_df["% Deviating"].round(1)
            st.dataframe(display_df, use_container_width=True, hide_index=True)

            deviating_only = display_df[display_df["Total Deviating"] > 0]
            if not deviating_only.empty:
                st.bar_chart(
                    deviating_only.set_index("Attribute")[["Value Mismatch", "Missing in AI Output"]]
                )

            st.download_button(
                label="Download Highlighted Comparison Report (.xlsx)",
                data=report_bytes,
                file_name="QC_Comparison_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.info("Upload both files to begin.")


else:  # ---------------- Generate Content ----------------
    st.markdown(
        f"""
        <div style="background-color:{NAVY};padding:18px 24px;border-radius:6px;margin-bottom:14px;">
            <h2 style="color:white;margin:0;">Category Content Generator</h2>
            <p style="color:#D6DCE5;margin:4px 0 0 0;">
                Choose a category, upload your file, and generate DN, LVN, PD, and
                Size & Fit content per the nomenclature rules. Results are written
                back into <code>productDisplayName</code>, <code>listViewName</code>,
                <code>Product Details</code>, and <code>sizeAndFitDescription</code>.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    category = st.selectbox("1. Category", options=CATEGORIES)
    content_file = st.file_uploader("2. Upload file with SKU attribute data", type=["xlsx", "xls", "csv"], key="content_upl")

    if content_file:
        sheets = load_sheets(content_file)
        sheet_name, df = pick_sheet("Data file", sheets, "content")
        df.columns = [str(c).strip() for c in df.columns]

        required_fields = get_required_fields(category)
        available_cols = ["(none)"] + list(df.columns)

        st.markdown("### 3. Confirm field mapping")
        st.caption(
            "Auto-matched where possible. Review and correct any that look wrong "
            "before generating — this drives every SKU's output."
        )

        mapping = {}
        cols_per_row = 3
        field_chunks = [required_fields[i:i + cols_per_row] for i in range(0, len(required_fields), cols_per_row)]
        for chunk in field_chunks:
            ui_cols = st.columns(cols_per_row)
            for i, field in enumerate(chunk):
                guess = guess_column(field, list(df.columns))
                default_idx = available_cols.index(guess) if guess in available_cols else 0
                with ui_cols[i]:
                    mapping[field] = st.selectbox(
                        field, options=available_cols, index=default_idx, key=f"map_{field}"
                    )
        mapping = {k: (None if v == "(none)" else v) for k, v in mapping.items()}

        st.markdown("### 4. Target columns (written back into your file)")
        target_map = {}
        tcol1, tcol2, tcol3, tcol4 = st.columns(4)
        for label, ui_col in zip(TARGET_COLUMNS, (tcol1, tcol2, tcol3, tcol4)):
            default_col = TARGET_COLUMNS[label]
            options = list(df.columns) if default_col in df.columns else list(df.columns) + [default_col]
            with ui_col:
                target_map[label] = st.selectbox(
                    f"{label} →", options=options,
                    index=options.index(default_col) if default_col in options else 0,
                    key=f"target_{label}",
                )

        apply_lvn_limit = st.checkbox("Enforce a max character length on LVN (truncate lowest-priority fields first)", value=False)
        lvn_max_length = None
        if apply_lvn_limit:
            lvn_max_length = st.number_input("Max LVN length", min_value=10, max_value=200, value=50, step=1)

        if st.button("Generate Content", type="primary"):
            with st.spinner("Generating content for every SKU..."):
                out_df = df.copy()
                for target_col in target_map.values():
                    if target_col not in out_df.columns:
                        out_df[target_col] = ""

                changed_cells = []  # (row_idx, target_col)
                for idx, row in df.iterrows():
                    result = generate_all(category, row, mapping, lvn_max_length=lvn_max_length)
                    for label, text in result.items():
                        target_col = target_map[label]
                        out_df.at[idx, target_col] = text
                        if text.strip():
                            changed_cells.append((idx, target_col))

                buf = io.BytesIO()
                out_df.to_excel(buf, index=False, sheet_name=sheet_name)
                buf.seek(0)

                # Re-open with openpyxl to highlight the generated cells green
                wb = load_workbook(buf)
                ws = wb[sheet_name]
                header = [c.value for c in ws[1]]
                col_idx = {name: i + 1 for i, name in enumerate(header)}
                fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
                for row_i, target_col in changed_cells:
                    if target_col in col_idx:
                        ws.cell(row=row_i + 2, column=col_idx[target_col]).fill = fill

                final_buf = io.BytesIO()
                wb.save(final_buf)
                final_buf.seek(0)
                report_bytes = final_buf.read()

            st.success(f"Generated content for {len(df)} SKU(s).")
            st.dataframe(
                out_df[[c for c in [mapping.get(f) for f in required_fields[:1]] if c] + list(target_map.values())].head(20),
                use_container_width=True,
            )

            st.download_button(
                label="Download Updated File (.xlsx)",
                data=report_bytes,
                file_name="Generated_Content.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.info("Upload a file to begin.")
